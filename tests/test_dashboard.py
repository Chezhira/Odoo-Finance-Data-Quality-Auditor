from __future__ import annotations

from io import BytesIO

from openpyxl import load_workbook
import pytest

from odoo_finance_data_auditor.dashboard import (
    apply_exception_filters,
    build_kpis,
    chart_height,
    count_by_dimension,
    friendly_source_model,
    integer_tick_values,
    load_uploaded_csv_exports,
    load_dashboard_results,
    match_uploaded_csv_files,
    missing_required_csv_files,
    required_csv_filenames,
    upload_status_rows,
    workbook_bytes,
)
from odoo_finance_data_auditor.loader import SchemaValidationError


def test_dashboard_results_load_registered_checks(sample_data_dir):
    exceptions, exception_rows = load_dashboard_results(sample_data_dir)

    assert len(exceptions) == 13
    assert len(exception_rows) == 13
    assert {"risk_level", "issue_type", "source_model", "recommended_action"}.issubset(exception_rows.columns)


def test_dashboard_kpis_count_current_sample_exceptions(sample_data_dir):
    _, exception_rows = load_dashboard_results(sample_data_dir)

    assert build_kpis(exception_rows) == {
        "total_checks": 10,
        "total_exceptions": 13,
        "high_risk_exceptions": 10,
        "exception_types": 10,
    }


def test_dashboard_filters_by_risk_issue_and_source(sample_data_dir):
    _, exception_rows = load_dashboard_results(sample_data_dir)

    filtered = apply_exception_filters(
        exception_rows,
        risk_levels=["medium"],
        issue_types=["Manual journal above configured threshold"],
        source_models=["journal_entries"],
    )

    assert filtered["record_id"].tolist() == ["JL008"]


def test_friendly_source_model_labels_known_and_fallback_values():
    assert friendly_source_model("journal_entries") == "Journal Entries"
    assert friendly_source_model("bank_statement_lines") == "Bank Statement Lines"
    assert friendly_source_model("custom_model_name") == "Custom Model Name"


def test_count_by_dimension_sorts_issue_types_and_formats_sources(sample_data_dir):
    _, exception_rows = load_dashboard_results(sample_data_dir)

    issue_counts = count_by_dimension(exception_rows, "issue_type", "Issue Type")
    source_counts = count_by_dimension(
        exception_rows,
        "source_model",
        "ERP Area",
        friendly_labels=True,
    )

    assert issue_counts.iloc[0]["exception_count"] == 2
    assert "Vendor Bills" in source_counts["ERP Area"].tolist()
    assert "Journal Entries" in source_counts["ERP Area"].tolist()


def test_chart_height_scales_with_bounds():
    assert chart_height(1) == 180
    assert chart_height(20) == 420


def test_integer_tick_values_cover_whole_number_counts():
    assert integer_tick_values(2) == [0, 1, 2]
    assert integer_tick_values(5) == [0, 1, 2, 3, 4, 5]


def test_dashboard_workbook_bytes_are_excel_file(sample_data_dir):
    _, exception_rows = load_dashboard_results(sample_data_dir)

    payload = workbook_bytes(exception_rows)

    assert payload[:2] == b"PK"
    workbook = load_workbook(filename=BytesIO(payload))
    assert workbook.sheetnames == ["Summary", "Exceptions"]


class NamedBytesIO(BytesIO):
    def __init__(self, name: str, payload: bytes):
        super().__init__(payload)
        self.name = name


def test_required_csv_filenames_match_expected_upload_contract():
    assert required_csv_filenames() == [
        "accounts.csv",
        "journal_entries.csv",
        "vendor_bills.csv",
        "customer_invoices.csv",
        "inventory_valuation.csv",
        "bank_statement_lines.csv",
    ]


def test_uploaded_file_matching_uses_required_filenames_only():
    uploaded_files = [
        NamedBytesIO("accounts.csv", b""),
        NamedBytesIO("notes.csv", b""),
        NamedBytesIO("JOURNAL_ENTRIES.csv", b""),
    ]

    matched = match_uploaded_csv_files(uploaded_files)

    assert set(matched) == {"accounts.csv", "journal_entries.csv"}
    assert missing_required_csv_files(matched) == [
        "vendor_bills.csv",
        "customer_invoices.csv",
        "inventory_valuation.csv",
        "bank_statement_lines.csv",
    ]


def test_upload_status_rows_marks_uploaded_and_missing_files():
    rows = upload_status_rows({"accounts.csv": object()})

    assert rows.loc[rows["required_file"].eq("accounts.csv"), "status"].item() == "Uploaded"
    assert rows.loc[rows["required_file"].eq("vendor_bills.csv"), "status"].item() == "Missing"


def test_load_uploaded_csv_exports_reads_complete_sample_set(sample_data_dir):
    uploaded_files = [
        NamedBytesIO(path.name, path.read_bytes())
        for path in sample_data_dir.glob("*.csv")
    ]

    data = load_uploaded_csv_exports(uploaded_files)

    assert set(data) == {
        "accounts",
        "journal_entries",
        "vendor_bills",
        "customer_invoices",
        "bank_statement_lines",
        "inventory_valuation",
    }


def test_load_uploaded_csv_exports_rejects_invalid_columns(sample_data_dir):
    uploaded_files = [
        NamedBytesIO(path.name, path.read_bytes())
        for path in sample_data_dir.glob("*.csv")
        if path.name != "accounts.csv"
    ]
    uploaded_files.append(NamedBytesIO("accounts.csv", b"account_id,code\n1000,1000\n"))

    with pytest.raises(SchemaValidationError, match="accounts.csv missing required columns"):
        load_uploaded_csv_exports(uploaded_files)
